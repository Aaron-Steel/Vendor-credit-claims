"""Excel export of a promotion in one of the three views (internal / sales / vendor).

Mirrors the column subsets of the original template's Promo Form / Sales Copy / Vendor
Copy tabs, one sheet per retailer block plus a summary.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEAD = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="2F5496")
TITLE = Font(bold=True, size=14)
BOLD = Font(bold=True)

# (header, attribute path, number format) per view
VIEWS = {
    "internal": [
        ("Code", "line.product_code", None),
        ("Description", "line.description", None),
        ("Retailer Buy ex", "line.retailer_buy_ex", "#,##0.00"),
        ("Supplier Support", "result.supplier_support", "#,##0.00"),
        ("MG Support", "result.mg_support", "#,##0.00"),
        ("Total Support", "result.total_support", "#,##0.00"),
        ("Std Margin", "result.std_margin", "0.0%"),
        ("Promo Margin", "result.promo_margin", "0.0%"),
        ("Expected Sales", "result.expected_sales", "#,##0"),
        ("Supplier Claim", "result.supplier_claim", "#,##0.00"),
        ("MG Claim", "result.mg_claim", "#,##0.00"),
        ("RRP", "line.rrp_inc", "#,##0.00"),
        ("% Off", "line.pct_off", "0.0%"),
        ("Promo Price", "result.rec_sale_inc", "#,##0.00"),
    ],
    "sales": [
        ("Code", "line.product_code", None),
        ("Description", "line.description", None),
        ("Retailer Buy ex", "line.retailer_buy_ex", "#,##0.00"),
        ("Total Support", "result.total_support", "#,##0.00"),
        ("Std Margin", "result.std_margin", "0.0%"),
        ("Promo Margin", "result.promo_margin", "0.0%"),
        ("Expected Sales", "result.expected_sales", "#,##0"),
        ("RRP", "line.rrp_inc", "#,##0.00"),
        ("% Off", "line.pct_off", "0.0%"),
        ("Promo Price", "result.rec_sale_inc", "#,##0.00"),
    ],
    "vendor": [
        ("Code", "line.product_code", None),
        ("Description", "line.description", None),
        ("Supplier Support", "result.supplier_support", "#,##0.00"),
        ("Expected Sales", "result.expected_sales", "#,##0"),
        ("Supplier Claim", "result.supplier_claim", "#,##0.00"),
        ("RRP", "line.rrp_inc", "#,##0.00"),
        ("% Off", "line.pct_off", "0.0%"),
        ("Promo Price", "result.rec_sale_inc", "#,##0.00"),
    ],
}


def _resolve(lv, path):
    obj, attr = path.split(".")
    return getattr(getattr(lv, obj), attr)


def build_workbook(pv, view: str) -> io.BytesIO:
    cols = VIEWS[view]
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    p = pv.promo

    ws["A1"] = f"{p.name or 'Promotion'} — {view.title()} Copy"
    ws["A1"].font = TITLE
    meta = [("Brand", p.brand), ("Claim #", p.claim_number),
            ("Dates", f"{p.start_date} to {p.end_date} ({pv.weeks:g} wks)")]
    if view != "vendor":
        meta.append(("Customer expected claim (AUD)", round(pv.customer_expected_aud, 2)))
    if view == "internal":
        meta.append(("MacGear absorbs (AUD)", round(pv.mg_total_aud, 2)))
    if view in ("internal", "vendor"):
        meta.append(("Supplier claim (AUD)", round(pv.supplier_total_aud, 2)))
        meta.append((f"Supplier claim (USD @ {p.aud_usd_rate})",
                     round(pv.supplier_total_usd, 2)))
    for i, (k, v) in enumerate(meta, start=3):
        ws.cell(i, 1, k).font = BOLD
        ws.cell(i, 2, v)

    for rv in pv.retailers:
        title = rv.pr.retailer_name[:31] or "Retailer"
        sh = wb.create_sheet(title=title)
        for c, (h, _, _) in enumerate(cols, start=1):
            cell = sh.cell(1, c, h)
            cell.font = HEAD
            cell.fill = HEAD_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for r, lv in enumerate(rv.lines, start=2):
            for c, (_, path, fmt) in enumerate(cols, start=1):
                val = _resolve(lv, path)
                cell = sh.cell(r, c, val)
                if fmt and isinstance(val, (int, float)):
                    cell.number_format = fmt
        # widths
        sh.column_dimensions["A"].width = 16
        sh.column_dimensions["B"].width = 42
        for col_letter in [chr(ord("C") + i) for i in range(len(cols) - 2)]:
            sh.column_dimensions[col_letter].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
