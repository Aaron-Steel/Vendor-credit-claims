"""One-off: extract seed data (retailers + product master) from the promo templates.

Reads the AU and NZ promo templates and writes combined, country-tagged seed:
  data/seed_retailers.json  - [{name, rebate, country}]
  data/seed_products.json   - [{code, description, brand, status, rrp_inc, channel_prices, country}]

Retailers and product codes overlap by name/code between countries (with different
pricing), so every row carries a `country` (AU/NZ) and the app keys on (name|code, country).

Re-run whenever a template's price/validation sheets change.
"""
import json
import os
import warnings

import openpyxl

warnings.simplefilter("ignore")  # silence the data-validation extension warning

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DATA_DIR = os.path.join(ROOT, "data")

# Per-country template config.
COUNTRIES = [
    {"code": "AU", "file": "AU_Promo Form_TEMPLATE.xlsx",
     "master_sheet": "Master Price Level"},
    {"code": "NZ", "file": "NZ_Promo Form_TEMPLATE.xlsx",
     "master_sheet": "Master Price Level NZ"},
]

# Master-sheet columns that are NOT a retailer channel price.
NON_CHANNEL = {"Source.Name", "Brand", "Code", "Description",
               "AU Status", "NZ Status", "Base (RRP Inc)", "RRP ex GST"}


def extract_retailers(wb, country):
    ws = wb["Data Validation"]
    retailers, seen = [], set()
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        name = str(name).strip()
        if name in seen:          # template sometimes lists a name twice; keep the first
            continue
        seen.add(name)
        rebate = ws.cell(r, 2).value
        retailers.append({"name": name, "country": country,
                          "rebate": float(rebate) if isinstance(rebate, (int, float)) else 0.0})
    return retailers


def extract_products(wb, country, master_sheet):
    ws = wb[master_sheet]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}
    status_header = "NZ Status" if country == "NZ" else "AU Status"
    channel_headers = [h for h in headers if h and h not in NON_CHANNEL]
    products, seen = [], set()
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, col["Code"]).value
        if not code:
            continue
        code = str(code).strip()
        if code in seen:
            continue
        seen.add(code)
        rrp = ws.cell(r, col["Base (RRP Inc)"]).value
        prices = {}
        for h in channel_headers:
            v = ws.cell(r, col[h]).value
            if isinstance(v, (int, float)):
                prices[h] = float(v)
        desc = ws.cell(r, col["Description"]).value
        brand = ws.cell(r, col["Brand"]).value
        status = ws.cell(r, col[status_header]).value if status_header in col else None
        products.append({
            "code": code,
            "country": country,
            "description": str(desc).strip() if desc else "",
            "brand": str(brand).strip() if brand else "",
            "status": str(status).strip() if status else "",
            "rrp_inc": float(rrp) if isinstance(rrp, (int, float)) else None,
            "channel_prices": prices,
        })
    return products


def main():
    os.makedirs(DATA_DIR, exist_ok=True)
    all_retailers, all_products = [], []
    for c in COUNTRIES:
        path = os.path.join(ROOT, c["file"])
        if not os.path.exists(path):
            print(f"  (skipped {c['code']} - {c['file']} not found)")
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        rets = extract_retailers(wb, c["code"])
        prods = extract_products(wb, c["code"], c["master_sheet"])
        all_retailers += rets
        all_products += prods
        print(f"  {c['code']}: {len(rets)} retailers, {len(prods)} products")
    with open(os.path.join(DATA_DIR, "seed_retailers.json"), "w", encoding="utf-8") as f:
        json.dump(all_retailers, f, indent=2, ensure_ascii=False)
    with open(os.path.join(DATA_DIR, "seed_products.json"), "w", encoding="utf-8") as f:
        json.dump(all_products, f, indent=2, ensure_ascii=False)
    print(f"Wrote {len(all_retailers)} retailers and {len(all_products)} products to {DATA_DIR}")


if __name__ == "__main__":
    main()
